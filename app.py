from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
import calendar
from zipfile import BadZipFile


getcontext().prec = 28

INPUT_LOANS_FILE = "loans.xlsx"
INPUT_SHEET_NAME = "Sheet1"
OUTPUT_EXCEL_FILE = "amortization_schedule.xlsx"
OUTPUT_FINAL_ROWS_FILE = "amortization_final_rows.xlsx"
COLUMN_NAME_MAP = {
    "loan_number": "loan_number",
    "periods_months": "periods_months",
    "projected_close_date": "projected_close_date",
    "interest_start_date": "interest_start_date",
    "first_payment_date": "first_payment_date",
    "cycle_day": "cycle_day",
    "annual_rate_percent": "annual_rate_percent",
    "loan_amount": "loan_amount",
    "monthly_payment": "monthly_payment",
}


@dataclass
class ScheduleRow:
    loan_number: str
    period: int
    payment_date: date | None
    days: int | None
    projected_close_date: date | None
    beginning_balance: Decimal
    daily_interest: Decimal | None
    interest: Decimal | None
    payment: Decimal | None
    principal: Decimal | None
    extra_interest: Decimal | None
    ending_balance: Decimal


def parse_date(value: str) -> date:
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value}")


def quantize_money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def add_months(base_date: date, months: int, cycle_day: int) -> date:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(cycle_day, last_day)
    return date(year, month, day)


def build_schedule(
    loan_number: str,
    periods_months: int,
    projected_close_date: date | None,
    interest_start_date: date,
    first_payment_date: date,
    cycle_day: int,
    annual_rate_percent: Decimal,
    loan_amount: Decimal,
    monthly_payment: Decimal,
) -> list[ScheduleRow]:
    rows: list[ScheduleRow] = []
    balance = quantize_money(loan_amount)

    rows.append(
        ScheduleRow(
            loan_number=loan_number,
            period=0,
            payment_date=interest_start_date,
            days=None,
            projected_close_date=projected_close_date,
            beginning_balance=balance,
            daily_interest=None,
            interest=None,
            payment=None,
            principal=None,
            extra_interest=None,
            ending_balance=balance,
        )
    )

    extra_interest: Decimal | None = None
    if projected_close_date:
        extra_days = (interest_start_date - projected_close_date).days
        if extra_days < 0:
            extra_days = 0
        extra_daily_interest = quantize_money(
            (annual_rate_percent / Decimal(100)) * balance / Decimal(365)
        )
        extra_interest = quantize_money(extra_daily_interest * Decimal(extra_days))
        rows[0].extra_interest = extra_interest

    previous_date = interest_start_date
    for period in range(1, periods_months + 1):
        if period == 1:
            payment_date = first_payment_date
        else:
            payment_date = add_months(first_payment_date, period - 1, cycle_day)

        days = (payment_date - previous_date).days
        daily_interest = quantize_money(
            (annual_rate_percent / Decimal(100)) * balance / Decimal(365)
        )
        interest = quantize_money(daily_interest * Decimal(days))
        if period == 1 and extra_interest is not None:
            interest = quantize_money(interest + extra_interest)

        payment = monthly_payment
        if period == 1 and extra_interest is not None:
            payment = payment + extra_interest
        if period == periods_months:
            payment = balance + interest
        elif payment > balance + interest:
            payment = balance + interest

        principal = payment - interest
        ending_balance = quantize_money(balance - principal)
        if ending_balance < Decimal("0.00"):
            ending_balance = Decimal("0.00")

        rows.append(
            ScheduleRow(
                loan_number=loan_number,
                period=period,
                payment_date=payment_date,
                days=days,
                projected_close_date=None,
                beginning_balance=balance,
                daily_interest=daily_interest,
                interest=interest,
                payment=payment,
                principal=principal,
                extra_interest=None,
                ending_balance=ending_balance,
            )
        )

        balance = ending_balance
        previous_date = payment_date

    return rows


def format_money(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:,.2f}"


def display_schedule(rows: list[ScheduleRow]) -> None:
    headers = [
        "Loan #",
        "Period",
        "Payment Date",
        "Days",
        "Projected Close Date",
        "Begin Balance",
        "Daily Interest",
        "Interest",
        "Payment",
        "Principal",
        "Extra Interest",
        "End Balance",
    ]
    widths = [8, 6, 12, 6, 18, 14, 14, 12, 12, 12, 14, 14]
    header_row = "  ".join(h.ljust(w) for h, w in zip(headers, widths))
    print(header_row)
    print("-" * len(header_row))

    for row in rows:
        values = [
            row.loan_number,
            str(row.period),
            row.payment_date.isoformat() if row.payment_date else "",
            "" if row.days is None else str(row.days),
            row.projected_close_date.isoformat() if row.projected_close_date else "",
            format_money(row.beginning_balance),
            format_money(row.daily_interest),
            format_money(row.interest),
            format_money(row.payment),
            format_money(row.principal),
            format_money(row.extra_interest),
            format_money(row.ending_balance),
        ]
        line = "  ".join(v.ljust(w) for v, w in zip(values, widths))
        print(line)


def load_loans(
    file_path: str, sheet_name: str, column_map: dict[str, str]
) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to read Excel files. Install with: pip install openpyxl"
        ) from exc

    lower_path = file_path.lower()
    if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xlsm")):
        raise ValueError("Input file must be an .xlsx or .xlsm Excel file.")

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError("Input file is not a valid Excel workbook.") from exc

    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return []

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    header_index = {name: idx for idx, name in enumerate(headers)}

    def cell_value(row_values: tuple[object, ...], key: str) -> object:
        column_name = column_map[key]
        index = header_index[column_name]
        return row_values[index]

    loans: list[dict[str, object]] = []
    for row_values in rows_iter:
        def as_text(value: object) -> str:
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, date):
                return value.isoformat()
            return str(value).strip()

        loans.append(
            {
                "loan_number": as_text(cell_value(row_values, "loan_number")),
                "periods_months": int(as_text(cell_value(row_values, "periods_months"))),
                "projected_close_date": parse_date(
                    as_text(cell_value(row_values, "projected_close_date"))
                ),
                "interest_start_date": parse_date(
                    as_text(cell_value(row_values, "interest_start_date"))
                ),
                "first_payment_date": parse_date(
                    as_text(cell_value(row_values, "first_payment_date"))
                ),
                "cycle_day": int(as_text(cell_value(row_values, "cycle_day"))),
                "annual_rate_percent": Decimal(
                    as_text(cell_value(row_values, "annual_rate_percent"))
                ),
                "loan_amount": Decimal(as_text(cell_value(row_values, "loan_amount"))),
                "monthly_payment": Decimal(
                    as_text(cell_value(row_values, "monthly_payment"))
                ),
            }
        )
    return loans


def export_schedule_excel(rows: list[ScheduleRow], file_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import numbers
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to export Excel files. Install with: pip install openpyxl"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amortization"

    headers = [
        "Loan #",
        "Period",
        "Payment Date",
        "Days",
        "Projected Close Date",
        "Begin Balance",
        "Daily Interest",
        "Interest",
        "Payment",
        "Principal",
        "Extra Interest",
        "End Balance",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row.loan_number,
                row.period,
                row.payment_date,
                row.days,
                row.projected_close_date,
                float(row.beginning_balance) if row.beginning_balance is not None else None,
                float(row.daily_interest) if row.daily_interest is not None else None,
                float(row.interest) if row.interest is not None else None,
                float(row.payment) if row.payment is not None else None,
                float(row.principal) if row.principal is not None else None,
                float(row.extra_interest) if row.extra_interest is not None else None,
                float(row.ending_balance) if row.ending_balance is not None else None,
            ]
        )

    for cell in sheet["A"][1:]:
        cell.number_format = numbers.FORMAT_TEXT

    for cell in sheet["C"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    for cell in sheet["E"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    for column in ["F", "G", "H", "I", "J", "K", "L"]:
        for cell in sheet[column][1:]:
            cell.number_format = numbers.FORMAT_NUMBER_00

    workbook.save(file_path)


def export_final_rows_excel(rows: list[ScheduleRow], file_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import numbers
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to export Excel files. Install with: pip install openpyxl"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Final Rows"

    headers = [
        "Loan #",
        "First Payment Date",
        "First Payment Days",
        "Projected Close Date",
        "First Begin Balance",
        "First Daily Interest",
        "First Interest",
        "First Payment",
        "First Principal",
        "First Extra Interest",
        "First End Balance",
        "Final Payment Date",
        "Final Payment Days",
        "Final Begin Balance",
        "Final Daily Interest",
        "Final Interest",
        "Final Payment",
        "Final Principal",
        "Final Extra Interest",
        "Final End Balance",
    ]
    sheet.append(headers)

    first_payment_by_loan: dict[str, ScheduleRow] = {}
    final_by_loan: dict[str, ScheduleRow] = {}
    for row in rows:
        final_by_loan[row.loan_number] = row
        if row.period == 1 and row.loan_number not in first_payment_by_loan:
            first_payment_by_loan[row.loan_number] = row

    def to_float(value: Decimal | None) -> float | None:
        if value is None:
            return None
        return float(value)

    for loan_number, final_row in final_by_loan.items():
        first_row = first_payment_by_loan.get(loan_number)
        sheet.append(
            [
                loan_number,
                first_row.payment_date if first_row else None,
                first_row.days if first_row else None,
                first_row.projected_close_date if first_row else None,
                to_float(first_row.beginning_balance) if first_row else None,
                to_float(first_row.daily_interest) if first_row else None,
                to_float(first_row.interest) if first_row else None,
                to_float(first_row.payment) if first_row else None,
                to_float(first_row.principal) if first_row else None,
                to_float(first_row.extra_interest) if first_row else None,
                to_float(first_row.ending_balance) if first_row else None,
                final_row.payment_date if final_row else None,
                final_row.days if final_row else None,
                to_float(final_row.beginning_balance) if final_row else None,
                to_float(final_row.daily_interest) if final_row else None,
                to_float(final_row.interest) if final_row else None,
                to_float(final_row.payment) if final_row else None,
                to_float(final_row.principal) if final_row else None,
                to_float(final_row.extra_interest) if final_row else None,
                to_float(final_row.ending_balance) if final_row else None,
            ]
        )

    for cell in sheet["A"][1:]:
        cell.number_format = numbers.FORMAT_TEXT

    for cell in sheet["B"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    for cell in sheet["D"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    for cell in sheet["L"][1:]:
        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    for column in ["E", "F", "G", "H", "I", "J", "K", "N", "O", "P", "Q", "R", "S", "T"]:
        for cell in sheet[column][1:]:
            cell.number_format = numbers.FORMAT_NUMBER_00

    workbook.save(file_path)


def main() -> None:
    loans = load_loans(INPUT_LOANS_FILE, INPUT_SHEET_NAME, COLUMN_NAME_MAP)
    all_rows: list[ScheduleRow] = []
    for loan in loans:
        rows = build_schedule(
            loan_number=loan["loan_number"],
            periods_months=loan["periods_months"],
            projected_close_date=loan["projected_close_date"],
            interest_start_date=loan["interest_start_date"],
            first_payment_date=loan["first_payment_date"],
            cycle_day=loan["cycle_day"],
            annual_rate_percent=loan["annual_rate_percent"],
            loan_amount=loan["loan_amount"],
            monthly_payment=loan["monthly_payment"],
        )
        all_rows.extend(rows)

    display_schedule(all_rows)
    export_schedule_excel(all_rows, OUTPUT_EXCEL_FILE)
    print(f"Excel file saved to {OUTPUT_EXCEL_FILE}")
    export_final_rows_excel(all_rows, OUTPUT_FINAL_ROWS_FILE)
    print(f"Excel file saved to {OUTPUT_FINAL_ROWS_FILE}")


if __name__ == "__main__":
    main()
